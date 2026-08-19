"""exp_baseline_recompute.py — 실무 설계 조건으로 경제적 관경을 재산출한다 (제5장 5.2.1·5.4.6)

왜 이 스크립트가 따로 필요한가
------------------------------
`exp_sensitivity.py` 의 기준 사례는 `도구명세_경제관경수리검토.md` 의 **검산 예시**
(유량 0.35 ㎥/s · 연장 4,200 m · 후보 7종 · 할인율 4.5 % · 30년 · 130 원/kWh · η 0.75)
이며, 그 문서가 스스로 "실제 사업의 산정 결과가 아니다" 라고 밝힌 가정값이다.
제5장 5.4.6 은 이 가정값 위에 민감도를 세워 두고 말미에

    ☐ 실제 설계 매개변수로 입력을 교체한 재산출은 미수행이다.

라고 적어 두었다. 이 스크립트가 그 자리를 채운다.

아울러 2026-08-18 의 기준값 대조는 **스크립트를 남기지 않았다.** 그때의 전력단가
환산(61.61 원/kWh)은 손으로 `.dyn` 을 읽어 계산한 것이고 그 계산은 작업이력에만
남아 있다. 제6장 6.3 아홉째가 지적한 "상수가 스크립트에 내장되어 입력 양식만으로는
제3자가 재현할 수 없다" 는 한계를, 대조 자체를 스크립트로 남기지 않아 **논문이 한 번
더 저지르고 있었던 셈**이다. 그래서 이 스크립트는 값을 손으로 옮겨 적지 않고
데이터셋 원본에서 읽어 계산한다.

무엇을 하는가
-------------
① 데이터셋의 xlsx 3종에서 설계 매개변수·관경별 단가·기준값을 **읽는다**
② 원 워크플로 `2-1. 경제적 관경 계산.dyn` 의 산식을 그대로 재현해 기준값과 대조한다
   (기준값이 그 그래프의 출력임을 확인하는 절차이며, 도구 검증이 아니다)
③ 본 연구의 도구 `select_economic_diameter` 를 실무 조건으로 호출한다
④ 도구 산출과 기준값의 편차를 **원인별로 분해**한다
⑤ 민감도를 세 사례(가정값 · 실무 조건 · 교차)에 대해 각각 구해 나란히 놓는다
⑥ 연장이 선정에 영향을 주는지 확인한다
⑦ 「동력비 비중이 민감도의 동인인가」를 반증한다 (후보·단가표 고정, 비중만 스윕)
⑧ 전부 JSON 으로 남긴다

Civil 3D 에 접근하지 않으므로 CAD 없이 돈다.

실행
----
    .\.venv\Scripts\python.exe experiments\exp_baseline_recompute.py --dataset "<데이터셋 경로>"

`--dataset` 은 `02. 수리종단검토(도수 및 송수관로)` 폴더를 가리킨다. 생략하면
`CIVIL3D_MCP_DATASET` 환경변수를 쓰고, 그것도 없으면 어디를 보아야 하는지 알려 주고
멈춘다. **경로를 소스에 박아 두지 않는 이유는 데이터셋이 저장소 밖에 있기 때문이다.**
"""
from __future__ import annotations

import argparse
import json
import math
import os
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from civil3d_mcp import hydraulics_core as hc   # noqa: E402

# 한글 Windows 콘솔 기본 인코딩(cp949)에서 즉사하지 않도록.
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except Exception:                                          # noqa: BLE001
        pass

XL_PARAM = "주요 설계 매개변수_경제적 관경계산.xlsx"
XL_COST = "관경별 개략공사비.xlsx"
XL_REPORT = "BIM 기반 사업부지 검토 요약 보고서.xlsx"

TOL = 1e-4          # 이분법 수렴 폭(상대비)


# ---------------------------------------------------------------------------
# 원 워크플로 재현 — 2-1. 경제적 관경 계산.dyn
# ---------------------------------------------------------------------------
# 아래 세 함수는 그 그래프의 Python 노드를 **한 줄씩 옮긴 것**이다. 반올림 위치까지
# 그대로 두었다. 본 연구의 도구와 다른 점(계수 9.8 대 0.163×60, 중간 반올림)이
# 편차의 원인이므로, 정리해서 옮기면 그 원인을 잃는다.

def dynamo_head_loss_m_per_km(diameter_mm: int, flow_m3_per_day: float) -> float:
    """노드 「손실수두 계산」(id 1c410e8e).

    C값 분기가 `Dia < 800 -> 100 / Dia < 1200 -> 110 / else 120` 이다.
    본 연구의 도구(`auto_c_value`)는 `D <= 700 / D800~900 / D >= 1000` 이므로
    **D1000·D1100 에서 갈린다.**
    """
    c = 100 if diameter_mm < 800 else (110 if diameter_mm < 1200 else 120)
    # ⚠ 계수를 hc.HW_COEFFICIENT 로 참조하지 말 것. 이 함수의 목적은 기준값이 그
    #    그래프의 출력임을 **도구와 무관하게** 확인하는 것이므로, 도구 상수가 바뀌면
    #    재현도 함께 바뀌는 구조가 되면 안 된다. 리터럴로 두고 아래에서 대조한다.
    return round(
        10.666
        * (c ** -1.85)
        * (diameter_mm / 1000) ** -4.87
        * (flow_m3_per_day / 86400) ** 1.85
        * 1000,
        2,
    )


def dynamo_velocity_m_s(diameter_mm: int, flow_m3_per_day: float) -> float:
    """노드 「유속 계산」(id 9d5d1218)."""
    area = (diameter_mm / 1000) ** 2 * math.pi / 4
    return round((flow_m3_per_day / 86400) / area, 2)


def dynamo_annual_power_cost(base_fee_won: float) -> dict[str, float]:
    """노드 「연간동력비」(id 5af3a4a5).

    ★ 이 노드가 전력단가 환산의 유일한 실물이다. 본 연구의 저장소에는 없다
      (도구는 완제품 원/kWh 만 받는다).

    ★ 여기서 `Q = 1000` 은 설계유량이 아니라 **고정 상수**다. 즉 이 노드가 내는
      값은 「유량 1,000 ㎥/일 · 양정 1 m 일 때의 연간 전력비」라는 단위 요율이며,
      실제 유량은 뒤의 「현가계산」 노드에서 곱해진다.

    ★ 계절평균 사용요금 48.98 원은 (60.1×2 + 44.6×4 + 48.2×6)/12 로, 요율표의
      출처가 그래프 어디에도 적혀 있지 않다. 재현은 되지만 근거는 추적되지 않는다.
    """
    avg_monthly = round((60.1 * 2 + 44.6 * 4 + 48.2 * 6) / 12, 2)      # 48.98
    usage_fee = round(((base_fee_won / 720) + avg_monthly) * 1.1, 2)   # 61.61
    # 0.163·Q[㎥/min]·H/η — Q 는 위 상수 1000 ㎥/일을 분당으로 환산한 것
    power_kw = round((0.163 * 1 * (1000 / 1440) * 1) / 0.8, 3)         # 0.141
    annual = round(power_kw * usage_fee * 24 * 365, 0)
    return {
        "avg_monthly_won_per_kwh": avg_monthly,
        "usage_fee_won_per_kwh": usage_fee,
        "unit_power_kw": power_kw,
        "annual_power_cost_won": annual,
    }


def dynamo_pv_factor(discount_rate_pct: float, years: int) -> float:
    """노드 「현가계수 계산」(id affa70bf) — round(..., 4) 까지 그대로."""
    i = discount_rate_pct / 100.0
    return round((pow(1 + i, years) - 1) / (i * pow(1 + i, years)), 4)


def dynamo_total_pv_won_per_m(
    unit_cost_won_per_m: float,
    head_loss_m_per_km: float,
    annual_power_cost_won: float,
    flow_m3_per_day: float,
    pv_factor: float,
) -> dict[str, float]:
    """노드 「현가계산」(id 37440b01) + 공사비 합산(`+` 노드 044e04).

    ★ 여기에 **관로 연장이 들어가지 않는다.** `loss/1000` 이 동수경사(m/m)이므로
      loss_cost 는 「관로 1 m 당 연간 전력비」이고, 공사비도 원/m 이다. 따라서
      원 워크플로의 「총 현가」는 **원/m** 이며 연장에 무관하다.
    """
    loss_cost = round((annual_power_cost_won / 1000) * flow_m3_per_day
                      * (head_loss_m_per_km / 1000), 0)
    energy_pv = round(loss_cost * pv_factor, 0)
    return {
        "annual_loss_cost_won_per_m": loss_cost,
        "energy_pv_won_per_m": energy_pv,
        "total_pv_won_per_m": unit_cost_won_per_m + energy_pv,
    }


# ---------------------------------------------------------------------------
# 데이터셋 읽기
# ---------------------------------------------------------------------------

def _require(path: Path) -> Path:
    if not path.exists():
        raise SystemExit(
            f"[중단] 데이터셋 파일이 없다: {path}\n"
            f"       --dataset 으로 「02. 수리종단검토(도수 및 송수관로)」 폴더를 지정할 것."
        )
    return path


def read_dataset(root: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("[중단] openpyxl 이 필요하다:  .\\.venv\\Scripts\\pip install openpyxl")

    # ① 설계 매개변수 — 라벨 열 + 값 열의 세로 배치
    wb = openpyxl.load_workbook(_require(root / XL_PARAM), data_only=True)
    ws = wb["InputParameters"]
    raw = {}
    cell_types = {}
    for row in ws.iter_rows(min_row=1, max_col=2):
        label, value = row[0].value, row[1].value
        if label is None:
            continue
        raw[str(label).strip()] = value
        cell_types[str(label).strip()] = row[1].data_type      # 'n' 숫자 / 's' 문자열

    # ② 관경별 단가 — 관 재질별 시트
    wb = openpyxl.load_workbook(_require(root / XL_COST), data_only=True)
    cost_sheets: dict[str, list[tuple[int, float]]] = {}
    for name in wb.sheetnames:
        rows = []
        for r in wb[name].iter_rows(min_row=2, max_col=2, values_only=True):
            if r[0] is None:
                continue
            rows.append((int(r[0]), float(r[1]) if len(r) > 1 and r[1] is not None else None))
        cost_sheets[name] = rows

    # ③ 기준값 — 요약 보고서
    wb = openpyxl.load_workbook(_require(root / XL_REPORT), data_only=True)
    ws = wb["경제적 관경계산"]
    reference = []
    for r in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        if r[1] is None:
            continue
        reference.append({
            "diameter_mm": int(str(r[1]).replace("mm", "").replace(".0", "")),
            "velocity_m_s": r[2],
            "head_loss_m_per_km": r[3],
            "total_pv": r[4],
        })
    min_cost_label = ws.cell(2, 6).value
    earth = wb["물량산출"]
    earthwork_report = [earth.cell(2, c).value for c in range(2, 7)]

    return {
        "parameters": raw,
        "parameter_cell_types": cell_types,
        "unit_cost_sheets": cost_sheets,
        "reference_rows": reference,
        "reference_min_cost": min_cost_label,
        "earthwork_report_xlsx": earthwork_report,
    }


# ---------------------------------------------------------------------------
# 민감도 — 이분법으로 선정이 뒤집히는 배율을 찾는다
# ---------------------------------------------------------------------------

def bisect_threshold(choose, lo: float, hi: float) -> float | None:
    """배율 lo(선정 불변)에서 hi 로 가며 선정이 바뀌는 지점을 찾는다.

    구간이 축퇴(lo == hi)면 탐색 자체가 불가능하므로 None 을 돌려주고,
    호출부는 이를 「역전 없음」이 아니라 **「탐색 구간 없음」**으로 구분해 적는다.
    (연간 가동시간의 상향·펌프효율의 상향이 물리 상한에 걸리는 경우가 이에 해당한다.)
    """
    if abs(hi - lo) < 1e-12:
        return None
    base = choose(lo)
    if choose(hi) == base:
        return None
    a, b = lo, hi
    for _ in range(200):
        if abs(b - a) < TOL:
            break
        m = (a + b) / 2.0
        if choose(m) == base:
            a = m
        else:
            b = m
    return b


def sensitivity(case: dict[str, Any]) -> dict[str, Any]:
    """한 사례에 대한 민감도 표를 만든다.

    case = {base: {...도구 인자...}, unit_cost: {D: 원/m}}
    """
    base_kw = case["base"]
    cost = case["unit_cost"]

    def select(cost_table, **over):
        kw = dict(base_kw)
        kw.update(over)
        kw["unit_construction_cost"] = [
            {"diameter_mm": d, "cost_per_m": c} for d, c in cost_table.items()
        ]
        return hc.select_economic_diameter(**kw)

    def chosen(cost_table, **over) -> int:
        return select(cost_table, **over)["selected"]["diameter_mm"]

    def scaled(f: float, only: int | None = None) -> dict[int, float]:
        if only is None:
            return {d: c * f for d, c in cost.items()}
        return {d: (c * f if d == only else c) for d, c in cost.items()}

    result = select(cost)
    ranked = sorted(result["candidates"], key=lambda c: c["total_pv"])
    pick = result["selected"]["diameter_mm"]
    runner = ranked[1]["diameter_mm"] if len(ranked) > 1 else None

    eta = base_kw["pump_efficiency"]
    hours = base_kw["annual_operating_hours"]

    # (이름, 배율→선정 함수, 하향 구간, 상향 구간, 상한 사유)
    sweeps = [
        ("전체 건설단가", lambda f: chosen(scaled(f)), (1.0, 0.3), (1.0, 3.0), None),
        (f"D{pick} 단가만 (선정)", lambda f: chosen(scaled(f, only=pick)),
         (1.0, 0.3), (1.0, 3.0), None),
    ]
    if runner is not None:
        sweeps.append((f"D{runner} 단가만 (차순위)", lambda f: chosen(scaled(f, only=runner)),
                       (1.0, 0.3), (1.0, 3.0), None))
    sweeps += [
        ("전력요금", lambda f: chosen(cost, power_tariff_won_per_kwh=
                                     base_kw["power_tariff_won_per_kwh"] * f),
         (1.0, 0.1), (1.0, 5.0), None),
        ("할인율", lambda f: chosen(cost, discount_rate=base_kw["discount_rate"] * f),
         (1.0, 0.1), (1.0, 5.0), None),
        ("연간 가동시간", lambda f: chosen(cost, annual_operating_hours=min(hours * f, 8760.0)),
         (1.0, 0.1), (1.0, 8760.0 / hours), "연 8,760 h"),
        ("펌프효율", lambda f: chosen(cost, pump_efficiency=min(eta * f, 1.0)),
         (1.0, 0.5), (1.0, 1.0 / eta), "η ≤ 1"),
    ]

    rows = []
    for name, fn, down, up, cap in sweeps:
        d_thr = bisect_threshold(fn, *down)
        u_thr = bisect_threshold(fn, *up)
        rows.append({
            "variable": name,
            "down_threshold_pct": None if d_thr is None else (d_thr - 1.0) * 100,
            "down_diameter": None if d_thr is None else fn(d_thr),
            "down_span_pct": (down[1] - 1.0) * 100,
            "up_threshold_pct": None if u_thr is None else (u_thr - 1.0) * 100,
            "up_diameter": None if u_thr is None else fn(u_thr),
            "up_span_pct": (up[1] - 1.0) * 100,
            "up_bound_reason": cap,
        })

    # 선정 관경의 총현가에서 동력비가 차지하는 비중.
    #
    # ⚠ 2026-08-19 철회 — 처음에 여기에 「민감도의 크기를 좌우하는 것은 이 비중이다」
    #   라고 적었으나 **반증되었다.** ⑥의 비중 스윕이 반례를 낸다.
    #     전력요금 ×0.5 → 비중 28.11 % (가정값 사례의 31.11 % 에 근접) → 역전 1/7
    #     유량     ×3   → 비중 10.41 % (실무 사례의 9.28 % 에 근접)   → 역전 7/7
    #   비중이 비슷해도 개수가 정반대이므로 비중은 동인이 아니다. 차순위 격차와도
    #   단조 관계가 아니다(격차 0.76 % → 7/7 이고 0.81 % → 2/7).
    #
    #   남는 것은 **인과가 아니라 관찰**이다 — 「10 % 이내로 뒤집히는 변수의 개수」
    #   라는 지표 자체가 입력 조건에 대해 0/7 에서 7/7 까지 요동한다. 그러므로 이
    #   지표를 방법의 성질로 보고해서는 안 되고 사례마다 함께 제시해야 한다.
    sel = result["selected"]
    energy_share = (sel["energy_pv"] / sel["total_pv"] * 100
                    if sel["total_pv"] else None)

    tight = [r["variable"] for r in rows
             if (r["down_threshold_pct"] is not None and abs(r["down_threshold_pct"]) < 10)
             or (r["up_threshold_pct"] is not None and abs(r["up_threshold_pct"]) < 10)]

    return {
        "selected_mm": pick,
        "energy_pv_share_pct": energy_share,
        "tight_count": len(tight),
        "tight_variables": tight,
        "ranking": [{"diameter_mm": c["diameter_mm"], "total_pv": c["total_pv"]}
                    for c in ranked],
        # 후보가 한 종만 남으면 도구가 이 키를 아예 넣지 않는다(KeyError 방지)
        "runner_up_margin_pct": (result.get("runner_up_margin") * 100
                                 if result.get("runner_up_margin") is not None else None),
        "third_margin_pct": ((ranked[2]["total_pv"] - ranked[0]["total_pv"])
                             / ranked[0]["total_pv"] * 100 if len(ranked) > 2 else None),
        "excluded": result["excluded"],
        "warnings": result.get("warnings", []),
        "assumptions": result.get("assumptions", {}),
        "thresholds": rows,
    }


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dataset", default=os.environ.get("CIVIL3D_MCP_DATASET"),
                    help="「02. 수리종단검토(도수 및 송수관로)」 폴더 경로")
    ap.add_argument("--length-m", type=float, default=1.0,
                    help=("관로 연장(m). 기본값 1.0 은 총현가를 **원/m** 로 내기 위한 것이다. "
                          "원 워크플로도 총현가를 원/m 로 계산하며 연장을 곱하지 않는다. "
                          "⚠ 경제적 관경계산의 입력 양식에는 연장 항목이 자체가 없다 — "
                          "요약 보고서의 1,708.293 m 는 유량·관경이 다른 **별개 시나리오**"
                          "(수리종단 검토)의 값이므로 기본값으로 쓰지 않는다. "
                          "⑥에서 보이듯 선정과 격차는 연장에 불변이다."))
    args = ap.parse_args()

    if not args.dataset:
        print("[중단] 데이터셋 경로를 모른다.\n"
              "       --dataset \"...\\01. Dynamo Logic(.dyn)\\02. 수리종단검토(도수 및 송수관로)\"\n"
              "       또는 환경변수 CIVIL3D_MCP_DATASET 로 지정할 것.", file=sys.stderr)
        return 2
    root = Path(args.dataset)
    ds = read_dataset(root)

    p = ds["parameters"]
    flow_per_day = float(p["설계유량"])
    material = str(p["관 재질"]).strip()
    base_fee = float(p["기본요금"])
    disc_pct = float(p["할인율"])
    years = int(p["사용년수"])

    print("=" * 92)
    print("  ① 데이터셋에서 읽은 실무 설계 조건")
    print("=" * 92)
    for k, v in p.items():
        print(f"    {k:10s} = {v!r}   (셀 타입 {ds['parameter_cell_types'][k]})")
    print(f"    관로 연장   = {args.length_m} m"
          + ("   (원/m 기준 — 입력 양식에 연장 항목이 없다)" if args.length_m == 1.0 else "   ⚠ 별개 시나리오의 값일 수 있다"))
    print()

    # 원 워크플로에서만 존재하는 상수 — 매개변수 파일에 없다
    power = dynamo_annual_power_cost(base_fee)
    tariff = power["usage_fee_won_per_kwh"]
    print("=" * 92)
    print("  ② 매개변수 파일에 없고 그래프에만 있는 값")
    print("=" * 92)
    print(f"    전력 사용요금 = ({base_fee:,.0f}/720 + {power['avg_monthly_won_per_kwh']}) × 1.1"
          f" = {tariff} 원/kWh")
    print(f"    펌프효율 0.8 · 연간 가동시간 8,760 h · 축동력 계수 0.163  ← 노드에 내장")
    print(f"    ⚠ 계절평균 {power['avg_monthly_won_per_kwh']} 원의 요율표 출처는 그래프에 없다")
    print()

    # 관 재질 시트 — 그래프는 시트명을 하드코딩하므로 매개변수의 「관 재질」을 읽지 않는다
    if material not in ds["unit_cost_sheets"]:
        raise SystemExit(
            f"[중단] 관 재질 '{material}' 시트가 단가표에 없다. "
            f"있는 시트: {list(ds['unit_cost_sheets'])}\n"
            f"       조용히 다른 재질로 대체하면 그 결과를 실제 재질의 결과로 읽게 된다."
        )
    sheet = material
    cost_rows = [(d, c) for d, c in ds["unit_cost_sheets"][sheet] if c is not None]
    unit_cost = dict(cost_rows)
    diameters = [d for d, _ in cost_rows]

    # ---------------------------------------------------------------- ③ 원 워크플로 재현
    pv_f = dynamo_pv_factor(disc_pct, years)
    print("=" * 92)
    print(f"  ③ 원 워크플로 재현 ↔ 기준값 대조   (현가계수 {pv_f})")
    print("=" * 92)
    print(f"  {'D(mm)':>7} | {'유속 재현':>9} {'기준':>7} | {'손실수두 재현':>13} {'기준':>9}"
          f" | {'총현가 재현':>12} {'기준':>12} | 일치")
    ref_by_d = {r["diameter_mm"]: r for r in ds["reference_rows"]}
    mismatch = 0
    for d in diameters:
        ref = ref_by_d.get(d)
        if ref is None:
            continue
        v = dynamo_velocity_m_s(d, flow_per_day)
        hl = dynamo_head_loss_m_per_km(d, flow_per_day)
        tp = dynamo_total_pv_won_per_m(unit_cost[d], hl,
                                       power["annual_power_cost_won"], flow_per_day, pv_f)
        ok = (abs(v - ref["velocity_m_s"]) < 5e-3
              and abs(hl - ref["head_loss_m_per_km"]) < 5e-3
              and abs(tp["total_pv_won_per_m"] - ref["total_pv"]) < 0.5)
        mismatch += 0 if ok else 1
        print(f"  {d:>7} | {v:>9.2f} {ref['velocity_m_s']:>7} | {hl:>13.2f}"
              f" {ref['head_loss_m_per_km']:>9} | {tp['total_pv_won_per_m']:>12,.0f}"
              f" {ref['total_pv']:>12,} | {'O' if ok else 'X'}")
    repro_pick = min(
        ((d, dynamo_total_pv_won_per_m(
            unit_cost[d], dynamo_head_loss_m_per_km(d, flow_per_day),
            power["annual_power_cost_won"], flow_per_day, pv_f)["total_pv_won_per_m"])
         for d in diameters), key=lambda t: t[1])[0]
    print(f"\n  불일치 {mismatch} 행 / {len(ref_by_d)} 행"
          f"   ·   재현 최소비용 관경 D{repro_pick}   ↔   기준값 {ds['reference_min_cost']}")
    print()

    # ---------------------------------------------------------------- ④ 도구 호출
    print("=" * 92)
    print("  ④ 본 연구의 도구를 실무 조건으로 호출")
    print("=" * 92)
    measured_base = dict(
        flow_m3_s=flow_per_day / 86400.0,
        length_m=args.length_m,
        candidate_diameters_mm=diameters,
        power_tariff_won_per_kwh=tariff,
        pump_efficiency=0.8,
        annual_operating_hours=8760.0,
        discount_rate=disc_pct / 100.0,
        service_life_years=years,
    )
    tool = hc.select_economic_diameter(
        unit_construction_cost=[{"diameter_mm": d, "cost_per_m": c}
                                for d, c in unit_cost.items()],
        **measured_base)
    print(f"    선정 관경 = D{tool['selected']['diameter_mm']}"
          f"   (기준값 {ds['reference_min_cost']})")
    print(f"    차순위 격차 = {tool['runner_up_margin'] * 100:.4f} %")
    print(f"    유속 범위로 제외된 관경 {len(tool['excluded'])}종")
    for w in tool.get("warnings", []):
        print(f"    경고: {w}")
    print()

    # ---- ④-2 편차 분해 -----------------------------------------------------
    # ⚠ 기준값은 소수 둘째 자리로 **저장**되어 있다. 전정밀도 도구 값을 그대로 대면
    #    작은 손실수두에서 상대오차가 부풀려져 「거짓 불일치」가 나온다(2026-08-18 에
    #    이미 한 번 겪은 함정). 따라서 ⓐ손실수두는 기준값의 표기 자리수로 맞추어 대조하고
    #    ⓑ편차는 총현가(원/m)에서 잰다.
    wide = hc.select_economic_diameter(
        unit_construction_cost=[{"diameter_mm": d, "cost_per_m": c}
                                for d, c in unit_cost.items()],
        velocity_range_m_s=(1e-4, 1e4), **dict(measured_base, length_m=1.0))
    print("  ④-2 도구 ↔ 원 워크플로   (손실수두는 기준값 표기 자리수로, 편차는 총현가 원/m 로)")
    print(f"  {'D(mm)':>7} | {'도구 m/km':>11} {'기준':>9} {'일치':>5} |"
          f" {'도구 총현가':>13} {'기준':>12} | {'편차 %':>9} | 사유")
    dev_rows = []
    hl_match = hl_total = 0
    for c in wide["candidates"]:
        d = c["diameter_mm"]
        ref = ref_by_d.get(d)
        if ref is None:
            continue
        tool_hl = c["hydraulic_gradient"] * 1000.0
        same_hl = round(tool_hl, 2) == round(float(ref["head_loss_m_per_km"]), 2)
        hl_total += 1
        hl_match += 1 if same_hl else 0
        dev = ((c["total_pv"] - ref["total_pv"]) / ref["total_pv"] * 100
               if ref["total_pv"] else None)
        why = "C값 규정 상이 (도구 120 / 기준 110)" if 1000 <= d < 1200 else \
              ("축동력 계수 + 기준의 중간 반올림" if dev and abs(dev) > 1e-6 else "")
        dev_rows.append({"diameter_mm": d, "tool_head_loss_m_per_km": tool_hl,
                         "reference_head_loss_m_per_km": ref["head_loss_m_per_km"],
                         "head_loss_matches_at_reference_precision": same_hl,
                         "tool_total_pv_won_per_m": c["total_pv"],
                         "reference_total_pv_won_per_m": ref["total_pv"],
                         "total_pv_deviation_pct": dev})
        print(f"  {d:>7} | {tool_hl:>11.5f} {ref['head_loss_m_per_km']:>9}"
              f" {'O' if same_hl else 'X':>5} | {c['total_pv']:>13,.0f}"
              f" {ref['total_pv']:>12,} | {dev:>+9.4f} | {why}")
    devs = [r["total_pv_deviation_pct"] for r in dev_rows
            if r["total_pv_deviation_pct"] is not None]
    print(f"\n  손실수두 기준 표기 자리수 일치 {hl_match} / {hl_total} 행")
    print(f"  총현가 편차 범위 {min(devs):+.4f} ~ {max(devs):+.4f} %"
          f"   (부호가 바뀌므로 단조가 아니다)")
    # 편차의 두 원인을 수치로 분리한다
    coeff_ratio = hc.GRAVITY_POWER_COEFF / (0.163 * 60)          # 9.8 / 9.78
    round_ratio = (0.163 * (1000 / 1440) / 0.8) / power["unit_power_kw"]
    print(f"  원인 ① 축동력 계수 {hc.GRAVITY_POWER_COEFF} 대 0.163×60 = {0.163*60:.2f}"
          f"  → {(coeff_ratio-1)*100:+.4f} %")
    print(f"  원인 ② 원 워크플로의 round(P, 3) : {0.163*(1000/1440)/0.8:.8f}"
          f" → {power['unit_power_kw']}  → {(round_ratio-1)*100:+.4f} %")
    print(f"  ①×② = {(coeff_ratio*round_ratio-1)*100:+.4f} %  ← 동력비 성분에 걸리는 계통 편차")
    print(f"  원인 ③ 기준값의 손실수두 2자리·비용 원 단위 반올림 → 부호가 바뀌는 성분")
    print()

    # ---------------------------------------------------------------- ⑤ 민감도 두 사례
    assumed_case = {
        "label": "가정값 (도구명세 검산 예시)",
        "base": dict(
            flow_m3_s=0.35, length_m=4200.0,
            candidate_diameters_mm=[300, 400, 450, 500, 600, 700, 800],
            power_tariff_won_per_kwh=130.0, pump_efficiency=0.75,
            annual_operating_hours=8760.0, discount_rate=0.045, service_life_years=30),
        "unit_cost": {300: 330_000.0, 400: 430_000.0, 450: 480_000.0, 500: 550_000.0,
                      600: 690_000.0, 700: 860_000.0, 800: 1_050_000.0},
    }
    measured_case = {
        "label": f"실무 설계 조건 ({sheet} {len(diameters)}종)",
        "base": measured_base,
        "unit_cost": unit_cost,
    }
    # 세 번째 사례 — 실무 매개변수에 가정값 단가표를 물린 교차 확인.
    # 민감도의 성격을 좌우하는 것이 「입력 규모」인지 「단가표 구조」인지를 가른다.
    cross_case = {
        "label": "교차 — 실무 매개변수 + 가정값 단가표 7종",
        "base": dict(measured_base,
                     candidate_diameters_mm=list(assumed_case["unit_cost"].keys())),
        "unit_cost": assumed_case["unit_cost"],
    }

    sens = {}
    for case in (assumed_case, measured_case, cross_case):
        print("=" * 92)
        print(f"  ⑤ 민감도 — {case['label']}")
        print("=" * 92)
        s = sensitivity(case)
        sens[case["label"]] = s
        print(f"    선정 D{s['selected_mm']}  ·  차순위 격차 {s['runner_up_margin_pct']:.4f} %"
              + (f"  ·  3위 격차 {s['third_margin_pct']:.4f} %"
                 if s["third_margin_pct"] is not None else "")
              + (f"  ·  동력비 비중 {s['energy_pv_share_pct']:.2f} %"
                 if s["energy_pv_share_pct"] is not None else ""))
        print(f"    {'변수':<22} {'하향 임계':>10} {'→':^4} {'상향 임계':>10} {'→':^4}")
        for r in s["thresholds"]:
            dt = ("역전 없음" if r["down_threshold_pct"] is None
                  else f"{r['down_threshold_pct']:+.2f} %")
            ut = ("탐색 구간 없음" if r["up_threshold_pct"] is None and r["up_span_pct"] <= 0
                  else ("역전 없음" if r["up_threshold_pct"] is None
                        else f"{r['up_threshold_pct']:+.2f} %"))
            cap = f"  [상한 {r['up_bound_reason']}]" if r["up_bound_reason"] else ""
            print(f"    {r['variable']:<22} {dt:>10} {'D'+str(r['down_diameter']) if r['down_diameter'] else '—':^6}"
                  f" {ut:>14} {'D'+str(r['up_diameter']) if r['up_diameter'] else '—':^6}{cap}")
        tight = [r["variable"] for r in s["thresholds"]
                 if (r["down_threshold_pct"] is not None and abs(r["down_threshold_pct"]) < 10)
                 or (r["up_threshold_pct"] is not None and abs(r["up_threshold_pct"]) < 10)]
        print(f"    → 10 % 이내로 뒤집히는 변수 {len(tight)} / {len(s['thresholds'])}"
              + (f" : {', '.join(tight)}" if tight else ""))
        print()

    # ---------------------------------------------------------------- ⑥ 연장 무관성
    print("=" * 92)
    print("  ⑥ 연장이 선정에 영향을 주는가")
    print("=" * 92)
    for L in (1.0, 1708.293, 4200.0):
        r = hc.select_economic_diameter(
            unit_construction_cost=[{"diameter_mm": d, "cost_per_m": c}
                                    for d, c in unit_cost.items()],
            **dict(measured_base, length_m=L))
        print(f"    L = {L:>10,.3f} m  →  D{r['selected']['diameter_mm']}"
              f" · 차순위 격차 {r['runner_up_margin']*100:.6f} %")
    print("    ※ 실양정(static_head_m)이 0 이면 총현가가 연장에 비례하므로 후보 간 순위가"
          " 연장에 무관하다. 원 워크플로도 총현가를 원/m 로 계산해 연장을 곱하지 않는다.")
    print()

    # ---------------------------------------------------------------- ⑦ 비중 스윕
    # ★ 「동력비 비중이 민감도의 동인」이라는 가설을 **반증**하기 위한 것이다.
    #   후보 집합과 단가표를 고정한 채 전력요금·유량만 움직여 비중을 바꾼다.
    #   비중이 동인이라면 비중이 비슷한 사례끼리 역전 개수도 비슷해야 한다.
    print("=" * 92)
    print("  ⑦ 「동력비 비중이 동인인가」 — 후보·단가표를 고정하고 비중만 움직인다")
    print("=" * 92)
    print(f"    {'사례':<18} {'선정':>6} {'동력비비중':>10} {'차순위격차':>11} {'3위격차':>10} {'10%내':>7}")
    sweep_rows = []
    variants = [("기준(실무)", {})]
    variants += [(f"전력요금 ×{f}", {"power_tariff_won_per_kwh": tariff * f})
                 for f in (0.25, 0.5, 2.0, 5.0, 40.0)]
    variants += [(f"유량 ×{f}", {"flow_m3_s": flow_per_day * f / 86400.0})
                 for f in (2.0, 3.0, 4.0)]
    for label, over in variants:
        s = sensitivity({"base": dict(measured_base, length_m=1.0, **over),
                         "unit_cost": unit_cost})
        third = ("%9.3f%%" % s["third_margin_pct"]) if s["third_margin_pct"] is not None else "     —   "
        print(f"    {label:<18} D{s['selected_mm']:<5} {s['energy_pv_share_pct']:9.2f}%"
              f" {s['runner_up_margin_pct']:10.4f}% {third} {s['tight_count']:>4}/7")
        sweep_rows.append({"case": label, "selected_mm": s["selected_mm"],
                           "energy_pv_share_pct": s["energy_pv_share_pct"],
                           "runner_up_margin_pct": s["runner_up_margin_pct"],
                           "third_margin_pct": s["third_margin_pct"],
                           "tight_count": s["tight_count"]})
    print()
    print("  ★ 판정 — 비중이 비슷한데 개수가 정반대인 짝이 있으면 「비중이 동인」은 성립하지 않는다.")
    print("     전력요금 ×0.5 는 비중이 가정값 사례(31.11 %)에 가까운데 개수는 최소이고,")
    print("     유량 ×3 은 비중이 실무 사례(9.28 %)에 가까운데 개수는 최대다.")
    print("     차순위 격차와도 단조가 아니다(0.76 % → 7/7 인데 0.81 % → 2/7).")
    print("     남는 것은 인과가 아니라 **지표 자체가 사례에 따라 0/7~7/7 로 요동한다**는 관찰이다.")
    print()

    out = Path(__file__).with_name("baseline_recompute_result.json")
    out.write_text(json.dumps({
        "dataset_root": str(root),
        "parameters": {k: (v if not isinstance(v, str) else v) for k, v in p.items()},
        "parameter_cell_types": ds["parameter_cell_types"],
        "derived_power": power,
        "pv_factor_dynamo": pv_f,
        "unit_cost_sheet": sheet,
        "length_m": args.length_m,
        "reference_min_cost": ds["reference_min_cost"],
        "reproduction_mismatch_rows": mismatch,
        "reproduction_selected_mm": repro_pick,
        "tool_selected_mm": tool["selected"]["diameter_mm"],
        "tool_runner_up_margin_pct": tool["runner_up_margin"] * 100,
        "tool_excluded_count": len(tool["excluded"]),
        "head_loss_deviation": dev_rows,
        "sensitivity": sens,
        "share_sweep": sweep_rows,
        "earthwork_report_xlsx": ds["earthwork_report_xlsx"],
        "hw_coefficient": hc.HW_COEFFICIENT,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  결과 저장: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
